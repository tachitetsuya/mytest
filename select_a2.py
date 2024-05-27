import openpyxl
import xlsxwriter

def select_a1_in_all_sheets(file_path):
    original_workbook = openpyxl.load_workbook(file_path)
    new_file_path = 'updated_' + file_path

    new_workbook = xlsxwriter.Workbook(new_file_path)

    for sheet_name in original_workbook.sheetnames:
        original_sheet = original_workbook[sheet_name]
        new_sheet = new_workbook.add_worksheet(sheet_name)
        
        for row_idx, row in enumerate(original_sheet.iter_rows()):
            for col_idx, cell in enumerate(row):
                new_sheet.write(row_idx, col_idx, cell.value)

        new_sheet.activate()
        new_sheet.select()
        new_sheet.set_selection('A1')

    new_workbook.close()

if __name__ == "__main__":
    excel_file_path = 'test2.xlsx'
    select_a1_in_all_sheets(excel_file_path)
